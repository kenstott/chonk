# Copyright (c) 2025 Kenneth Stott. MIT License.
# Canary: 91e47654-77df-45d9-b563-dd2bda42c754
#
# NOTICE: Use of this software for training artificial intelligence or
# machine learning models is strictly prohibited without explicit written
# permission from the copyright holder.

"""XLSX text extractor using openpyxl."""

from __future__ import annotations

from io import BytesIO
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from chonk.models import DocumentChunk

_openpyxl = None
try:
    import openpyxl as _openpyxl

    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False


def _extract_xlsx_content(wb: Any) -> str:  # noqa: ANN401
    """Extract text content from an openpyxl Workbook object."""
    sheets = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = []

        for row in sheet.iter_rows():
            cells = []
            for cell in row:
                if cell.value is not None:
                    cells.append(str(cell.value))
                else:
                    cells.append("")
            if any(c.strip() for c in cells):
                rows.append(" | ".join(cells))

        if rows:
            sheets.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))

    return "\n\n".join(sheets)


class XlsxExtractor:
    """Extract plain text from XLSX bytes."""

    def can_handle(self, doc_type: str) -> bool:
        return doc_type == "xlsx"

    def extract(self, data: bytes, source_path: str | None = None) -> str:
        if not _OPENPYXL_AVAILABLE:
            raise ImportError("pip install chonk[xlsx]")
        assert _openpyxl is not None

        wb = _openpyxl.load_workbook(BytesIO(data), data_only=True)
        return _extract_xlsx_content(wb)

    def annotate(
        self,
        chunks: list[DocumentChunk],
        data: bytes,
        source_path: str | None = None,
    ) -> list[DocumentChunk]:
        if not _OPENPYXL_AVAILABLE:
            raise ImportError("pip install chonk[xlsx]")
        assert _openpyxl is not None

        wb = _openpyxl.load_workbook(BytesIO(data), data_only=True)

        # Build map: row_text → (sheet_name, row_number_1indexed)
        row_map: list[tuple[str, str, int]] = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row_idx, row in enumerate(sheet.iter_rows(), 1):
                cells = [str(cell.value) if cell.value is not None else "" for cell in row]
                if any(c.strip() for c in cells):
                    row_map.append((sheet_name, " | ".join(cells), row_idx))

        for chunk in chunks:
            content = chunk.content
            matched_sheets: list[str] = []
            matched_rows: list[int] = []
            for sheet_name, row_text, row_idx in row_map:
                if row_text in content:
                    matched_sheets.append(sheet_name)
                    matched_rows.append(row_idx)
            if matched_rows:
                sheet = matched_sheets[0]
                chunk.source_detail = {
                    "sheet": sheet,
                    "row_start": min(matched_rows),
                    "row_end": max(matched_rows),
                }

        return chunks
